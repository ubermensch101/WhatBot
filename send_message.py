import time
from openpyxl import load_workbook

from docroom import WhatsApp


def send_common_message(file_path, message):

    messenger = WhatsApp()

    wb=load_workbook(file_path)
    ws=wb.active

    mobile_number_list=ws['A']

    for mobile_number_cell in mobile_number_list:

        current_val = mobile_number_cell.value

        if type(current_val) == str:
            current_val = current_val.replace(" ", "").replace("-", "").replace("+", "").replace(",", "")
            if current_val.isdigit():
                current_val = int(current_val)

        if type(current_val) == int or type(current_val) == float:

            if len(str(int(current_val))) == 10:
                current_mobile = "91"+str(int(current_val))
            else:
                current_mobile = str(int(current_val))

            messenger.send_message1(current_mobile, message)

    time.sleep(5)


def send_common_image_message(file_path, image_file_path, message):

    messenger = WhatsApp()

    wb = load_workbook(file_path)
    ws = wb.active

    mobile_number_list = ws['A']

    for mobile_number_cell in mobile_number_list:

        current_val = mobile_number_cell.value

        if type(current_val) == str:
            current_val = current_val.replace(" ", "").replace("-", "").replace("+", "").replace(",", "")
            if current_val.isdigit():
                current_val = int(current_val)

        if type(current_val) == int or type(current_val) == float:

            if len(str(int(current_val))) == 10:
                current_mobile = "91" + str(int(current_val))
            else:
                current_mobile = str(int(current_val))

            messenger.send_picture1(current_mobile, image_file_path, message)

    time.sleep(5)
